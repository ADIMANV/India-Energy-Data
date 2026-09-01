"""CEA installed capacity by state, sector and fuel.

    python -m gridscrapers.capacity load            # newest published month
    python -m gridscrapers.capacity load --month 2026-07

CEA publishes one workbook a month at a predictable path:

    cea.nic.in/wp-content/uploads/installed/{YYYY}/{MM}/IC_{Month}{YYYY}.xlsx

The state section (below the regional summary) repeats a fixed four-row block
per state — State / Private / Central / Sub-Total — against fuel columns. We
keep the three ownership rows and drop Sub-Total, which is derivable.

This is the only usable public source for a public/private split: the open
plant registry carries an owner for ~4% of Indian capacity. It is *capacity*,
not generation — see docs/DATA_GAPS.md §5 before deriving a share from it.

Two accounting notes, both easy to get wrong:

* The state section is captioned "including allocated shares in joint & central
  sector utilities", so a central station's capacity is apportioned to the
  states holding its PPAs. Summing these rows therefore does not reproduce the
  workbook's own regional summary sector-for-sector — 2026-07 gives
  private/state/central of roughly 55/25/21% here against 54/21/23% nationally.
  Neither is wrong; they answer different questions. Don't mix the two sections.
* NLC and DVC are central undertakings that CEA lists in the state table. They
  are not zones and are skipped, so state rows will not sum to the all-India
  total.
"""

import argparse
import datetime as dt
import io
import sys

import psycopg

from .db import archive_raw, get_dsn, insert_datapoints  # noqa: F401  (archive_raw)
from .http import make_client, request_raw

SOURCE = "cea_capacity"
PARSER_VERSION = 1
URL = ("https://cea.nic.in/wp-content/uploads/installed/"
       "{y}/{m:02d}/IC_{month_name}{y}.xlsx")

# Column offsets within the state section, 1-based worksheet columns.
COL_STATE, COL_SECTOR = 2, 3
FUEL_COLS = {
    "coal": 4, "lignite": 5, "gas": 6, "diesel": 7,
    "nuclear": 9, "hydro": 10, "res": 11,
}
SECTORS = {"state": "state", "private": "private", "central": "central"}

# CEA's own state labels -> our zone ids. Entries mapping to None are real rows
# in the workbook that are deliberately not zones: regional subtotals, and two
# central undertakings (NLC, DVC) that CEA lists alongside states.
STATE_ZONES = {
    "delhi": "IN-DL", "haryana": "IN-HR", "himachal pradesh": "IN-HP",
    "jammu & kashmir and ladakh": "IN-JK", "punjab": "IN-PB",
    "rajasthan": "IN-RJ", "uttar pradesh": "IN-UP", "uttarakhand": "IN-UK",
    "chandigarh": "IN-CH", "goa": "IN-GA", "gujarat": "IN-GJ",
    "madhya pradesh": "IN-MP", "chhattisgarh": "IN-CG", "maharashtra": "IN-MH",
    "andhra pradesh": "IN-AP", "telangana": "IN-TS", "karnataka": "IN-KA",
    "kerala": "IN-KL", "tamil nadu": "IN-TN", "puducherry": "IN-PY",
    "bihar": "IN-BR", "jharkhand": "IN-JH", "west bengal": "IN-WB",
    "odisha": "IN-OD", "sikkim": "IN-SK", "assam": "IN-AS",
    "arunachal pradesh": "IN-AR", "meghalaya": "IN-ML", "tripura": "IN-TR",
    "manipur": "IN-MN", "nagaland": "IN-NL", "mizoram": "IN-MZ",
}


def report_url(day: dt.date) -> str:
    return URL.format(y=day.year, m=day.month, month_name=day.strftime("%B"))


def fetch(day: dt.date | None = None):
    """Fetch a month's workbook. Walks back if the newest isn't published yet."""
    start = day or dt.date.today().replace(day=1)
    raws = []
    with make_client(legacy_tls=True) as client:
        for back in range(0, 4):
            y, m = divmod((start.year * 12 + start.month - 1) - back, 12)
            candidate = dt.date(y, m + 1, 1)
            raw = request_raw(client, SOURCE, "GET", report_url(candidate),
                              meta={"month": candidate.isoformat()})
            raws.append(raw)
            if (raw.http_status or 0) == 200 and raw.body[:2] == b"PK":
                break
            if day:  # explicit month: don't walk back past what was asked for
                break
    return raws


def _num(v) -> float | None:
    try:
        f = float(v)
    except (TypeError, ValueError):
        return None
    return f if f == f else None  # drop NaN


def parse(raw) -> list[tuple]:
    """-> [(as_of, zone, sector, fuel, capacity_mw)]. Empty on a bad fetch."""
    if not raw.body or (raw.http_status or 0) != 200 or raw.body[:2] != b"PK":
        return []
    import openpyxl

    ws = openpyxl.load_workbook(io.BytesIO(bytes(raw.body)), data_only=True).active

    # "(As on 31.07.2026)" — the report's own date, not our ingest time.
    as_of = None
    for row in ws.iter_rows(min_row=1, max_row=8, values_only=True):
        for cell in row:
            text = str(cell or "")
            if "as on" in text.lower():
                digits = "".join(c if c.isdigit() else " " for c in text).split()
                if len(digits) >= 3:
                    d, m, y = digits[0], digits[1], digits[2]
                    try:
                        as_of = dt.date(int(y), int(m), int(d))
                    except ValueError:
                        pass
    if as_of is None:
        as_of = dt.date.fromisoformat(raw.meta["month"])

    out, zone = [], None
    for i in range(1, ws.max_row + 1):
        label = " ".join(str(ws.cell(i, COL_STATE).value or "").split())
        sector = str(ws.cell(i, COL_SECTOR).value or "").strip().lower()
        if label:
            # A new labelled row starts a block; unknown labels (regional
            # subtotals, NLC/DVC) set zone to None so their rows are skipped
            # rather than silently attributed to the previous state.
            zone = STATE_ZONES.get(label.lower())
        if zone is None or sector not in SECTORS:
            continue
        for fuel, col in FUEL_COLS.items():
            mw = _num(ws.cell(i, col).value)
            if mw is not None:
                out.append((as_of, zone, SECTORS[sector], fuel, mw))
    return out


def load(day: dt.date | None = None) -> int:
    raws = fetch(day)
    rows: list[tuple] = []
    used = None
    for raw in raws:
        rows = parse(raw)
        if rows:
            used = raw
            break
    if not rows:
        tried = ", ".join(r.meta["month"] for r in raws)
        print(f"[{SOURCE}] no workbook parsed (tried {tried})", file=sys.stderr)
        return 0

    with psycopg.connect(get_dsn()) as conn:
        raw_id = archive_raw(conn, used)
        with conn.cursor() as cur:
            cur.executemany(
                """INSERT INTO cea_installed_capacity
                       (as_of, zone, sector, fuel, capacity_mw, raw_id)
                   VALUES (%s,%s,%s,%s,%s,%s)
                   ON CONFLICT (as_of, zone, sector, fuel)
                   DO UPDATE SET capacity_mw = EXCLUDED.capacity_mw,
                                 raw_id = EXCLUDED.raw_id,
                                 loaded_at = now()""",
                [r + (raw_id,) for r in rows],
            )
        conn.commit()
    print(f"[{SOURCE}] {len(rows)} rows, as_of {rows[0][0]}, "
          f"{len({r[1] for r in rows})} zones", file=sys.stderr)
    return len(rows)


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("command", choices=["load"])
    ap.add_argument("--month", help="YYYY-MM (default: newest published)")
    args = ap.parse_args()
    day = dt.date.fromisoformat(args.month + "-01") if args.month else None
    return 0 if load(day) else 1


if __name__ == "__main__":
    sys.exit(main())
